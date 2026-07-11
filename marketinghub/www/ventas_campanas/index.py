import frappe
from marketinghub.marketinghub.doctype.configuracion_meta.configuracion_meta import (
    get_meta_token as _conf_get_meta_token,
    get_meta_ad_account_id as _conf_get_meta_ad_account_id,
)
import requests
import json
from frappe.utils import today, add_days, flt, getdate


def get_context(context):
    if frappe.session.user == "Guest":
        frappe.local.flags.redirect_location = "/login?redirect-to=/ventas_campanas"
        raise frappe.Redirect
    user_roles = frappe.get_roles()
    allowed = ("Ventahub-Marketing-Ver", "Ventahub-Marketing-Administrar")
    if frappe.session.user != "Administrator" and not any(r in user_roles for r in allowed):
        frappe.throw("No tiene permisos para acceder a esta pagina.", frappe.PermissionError)
    context.no_cache = 1
    context.title = "Ventas & Campanas"


@frappe.whitelist()
def get_ventas_data(from_date=None, to_date=None, account_id=None):
    """Retorna datos de ventas enriquecidos con info de campana y contacto."""
    if not from_date:
        from_date = str(add_days(today(), -30))
    if not to_date:
        to_date = str(today())

    # 1. Facturas submitted en rango.
    # Nota: se derivo `custom_conversacion_chatwoot` originalmente pero en
    # Tiranidos no existe. Ahora se obtiene la conversacion via el Customer
    # → Lead (matching phone) → Conversacion Chatwoot mas reciente.
    placeholders_query = """
        SELECT
            si.name AS invoice_name,
            si.posting_date,
            si.grand_total,
            si.customer_name,
            si.customer,
            si.custom_anuncio AS invoice_anuncio,
            MIN(sii.sales_order) AS sales_order,
            MIN(so.owner) AS so_owner,
            MIN(so.custom_anuncio) AS so_anuncio
        FROM `tabSales Invoice` si
        LEFT JOIN `tabSales Invoice Item` sii ON sii.parent = si.name
        LEFT JOIN `tabSales Order` so ON so.name = sii.sales_order
        WHERE si.docstatus = 1
          AND si.posting_date BETWEEN %s AND %s
        GROUP BY si.name
        ORDER BY si.posting_date DESC
    """
    try:
        rows = frappe.db.sql(placeholders_query, (from_date, to_date), as_dict=True)
    except Exception as e:
        frappe.log_error(str(e), "ventas_campanas get_ventas_data SQL")
        return []

    if not rows:
        return []

    # Colecciones unicas.
    so_names = list(set(r.sales_order for r in rows if r.sales_order))
    customers = list(set(r.customer for r in rows if r.customer))
    owners = list(set(r.so_owner for r in rows if r.so_owner))

    # Derivar Conversacion Chatwoot por SI usando Customer → Lead (telefono).
    # Cada SI queda vinculado a la conversacion mas reciente del Lead del Customer.
    si_to_conv = {}
    if customers:
        customer_phones = frappe.get_all(
            "Customer", filters={"name": ["in", customers]},
            fields=["name", "custom_numero", "mobile_no"],
        )
        phone_to_customer = {}
        for c in customer_phones:
            phone = (c.get("custom_numero") or c.get("mobile_no") or "").strip()
            if phone:
                phone_to_customer[phone] = c.name

        customer_to_conv = {}
        if phone_to_customer:
            phones_list = list(phone_to_customer.keys())
            # Match exacto y match por ultimos 9 digitos.
            leads_exact = frappe.db.sql("""
                SELECT L.name AS lead_name, L.mobile_no,
                       (SELECT CC.name FROM `tabConversacion Chatwoot` CC
                        WHERE CC.lead = L.name ORDER BY CC.modified DESC LIMIT 1) AS conv_name
                FROM `tabLead` L
                WHERE L.mobile_no IN %(phones)s
            """, {"phones": phones_list}, as_dict=True)
            for row in leads_exact:
                cust_name = phone_to_customer.get((row.mobile_no or "").strip())
                if cust_name and row.conv_name:
                    customer_to_conv[cust_name] = row.conv_name

        for r in rows:
            conv = customer_to_conv.get(r.customer)
            if conv:
                si_to_conv[r.invoice_name] = conv

    conv_names = list(set(si_to_conv.values()))

    # 2. Nombres de usuarios (vendedoras)
    user_map = {}
    if owners:
        users = frappe.get_all("User", filters={"name": ["in", owners]}, fields=["name", "full_name"])
        user_map = {u.name: u.full_name for u in users}

    # 3. Telefonos de clientes
    phone_map = {}
    if customers:
        custs = frappe.get_all("Customer", filters={"name": ["in", customers]}, fields=["name", "custom_numero"])
        phone_map = {c.name: c.custom_numero for c in custs}

    # 4. Fecha de contacto e inbox (desde Conversacion Chatwoot)
    conv_date_map = {}
    conv_inbox_map = {}
    if conv_names and frappe.db.exists("DocType", "Conversacion Chatwoot"):
        convs = frappe.get_all("Conversacion Chatwoot",
            filters={"name": ["in", conv_names]},
            fields=["name", "creation", "inbox_name"])
        for c in convs:
            conv_date_map[c.name] = c.creation
            conv_inbox_map[c.name] = c.inbox_name

    # 5. Etiquetas lead por SO (fieldname del child es `etiqueta`, no `nombre`).
    etiquetas_map = {}
    if so_names:
        etiquetas = frappe.get_all("Etiqueta Lead Asociado",
            filters={"parent": ["in", so_names], "parenttype": "Sales Order"},
            fields=["parent", "etiqueta"])
        for e in etiquetas:
            etiquetas_map.setdefault(e.parent, []).append(e.etiqueta or "")

    # 6. WhatsApp empresa (numeros) por SO
    numeros_map = {}
    if so_names:
        numeros = frappe.get_all("Numeros Asociados",
            filters={"parent": ["in", so_names], "parenttype": "Sales Order"},
            fields=["parent", "numero"])
        for n in numeros:
            numeros_map.setdefault(n.parent, []).append(n.numero or "")

    # 7. Costos de productos y envio por SO
    costo_map, envio_map = _calc_costs_per_so(so_names)

    # 8. Cadena: etiqueta -> Anuncio -> Conjunto -> Campana
    all_etiquetas = set()
    for tags in etiquetas_map.values():
        all_etiquetas.update(tags)

    campaign_chain = _build_campaign_chain(all_etiquetas)

    # 9. Recopilar todas las fechas de contacto para expandir rango de gasto
    todas_fechas_contacto = set()
    for r in rows:
        conv = si_to_conv.get(r.invoice_name)
        if conv:
            fc = conv_date_map.get(conv)
            if fc:
                todas_fechas_contacto.add(str(fc)[:10])

    # Rango de gasto: desde la fecha de contacto mas antigua hasta la mas reciente
    spend_from = from_date
    spend_to = to_date
    if todas_fechas_contacto:
        fecha_min = min(todas_fechas_contacto)
        fecha_max = max(todas_fechas_contacto)
        if fecha_min < str(from_date):
            spend_from = fecha_min
        if fecha_max > str(to_date):
            spend_to = fecha_max

    # Gasto diario desde Meta API (a nivel anuncio, rango expandido, todos los meta_ids)
    spend_data = {}
    if account_id:
        unique_ad_meta_ids = set()
        for v in campaign_chain.values():
            for mid in v.get("all_meta_ids", []):
                if mid:
                    unique_ad_meta_ids.add(mid)
        if unique_ad_meta_ids:
            spend_data = _fetch_daily_spend(account_id, spend_from, spend_to, unique_ad_meta_ids)

    # 10. Pre-calcular datos por fila (anuncio_meta_id, fecha_contacto, gasto_dia_total)
    filas_pre = []
    for r in rows:
        so = r.sales_order
        etiquetas = etiquetas_map.get(so, [])

        campana_nombre = ""
        conjunto_nombre = ""
        anuncio_nombre = ""
        anuncio_meta_id = ""
        etiqueta_usada = ""
        codigo_campana = ", ".join(etiquetas)

        # Prioridad 1: atribucion CTWA directa (Sales Invoice.custom_anuncio o
        # Sales Order.custom_anuncio, seteados por el hook al submit del SI).
        anuncio_directo = r.get("invoice_anuncio") or r.get("so_anuncio")
        if anuncio_directo:
            chain_directo = _resolve_chain_from_anuncio(anuncio_directo)
            if chain_directo:
                campana_nombre = chain_directo["campana"]
                conjunto_nombre = chain_directo["conjunto_nombre"]
                anuncio_nombre = chain_directo["anuncio"]
                anuncio_meta_id = chain_directo["anuncio_meta_id"]

        # Prioridad 2: atribucion por etiqueta del Sales Order (legacy).
        if not anuncio_nombre:
            for tag in etiquetas:
                chain = campaign_chain.get(tag)
                if chain:
                    campana_nombre = chain["campana"]
                    conjunto_nombre = chain["conjunto_nombre"]
                    anuncio_nombre = chain["anuncio"]
                    anuncio_meta_id = chain["anuncio_meta_id"]
                    etiqueta_usada = tag
                    break

        conv = si_to_conv.get(r.invoice_name)
        fecha_contacto = ""
        if conv:
            fc = conv_date_map.get(conv)
            if fc:
                fecha_contacto = str(fc)[:10]

        whatsapp_empresa = ", ".join(numeros_map.get(so, [])) if so else ""
        if not whatsapp_empresa and conv:
            whatsapp_empresa = conv_inbox_map.get(conv, "")

        # Sumar gasto de TODOS los meta_ids de la misma etiqueta (cubre duplicados)
        gasto_dia_total = 0
        all_meta_ids = []
        if fecha_contacto:
            chain_data = campaign_chain.get(next((tag for tag in etiquetas if tag in campaign_chain), ""), {})
            all_meta_ids = chain_data.get("all_meta_ids", [])
            for mid in all_meta_ids:
                ad_spend = spend_data.get(mid, {})
                gasto = ad_spend.get(fecha_contacto, 0)
                gasto_dia_total += gasto

        filas_pre.append({
            "row": r, "so": so, "campana_nombre": campana_nombre,
            "conjunto_nombre": conjunto_nombre, "anuncio_nombre": anuncio_nombre,
            "anuncio_meta_id": anuncio_meta_id, "etiqueta_usada": etiqueta_usada,
            "codigo_campana": codigo_campana,
            "fecha_contacto": fecha_contacto, "whatsapp_empresa": whatsapp_empresa,
            "gasto_dia_total": gasto_dia_total
        })

    # 11. Contar TODAS las ventas historicas por etiqueta+fecha para prorratear gasto
    #     Agrupa por etiqueta (no meta_id) para cubrir duplicados del mismo anuncio
    from collections import Counter

    # Recopilar etiquetas+fecha que tienen gasto
    etiq_fecha_con_gasto = set()
    for fp in filas_pre:
        if fp["etiqueta_usada"] and fp["fecha_contacto"] and fp["gasto_dia_total"] > 0:
            etiq_fecha_con_gasto.add((fp["etiqueta_usada"], fp["fecha_contacto"]))

    conteo_etiq_fecha = Counter()
    if etiq_fecha_con_gasto:
        etiquetas_relevantes = list(set(ef[0] for ef in etiq_fecha_con_gasto))

        # Buscar TODAS las facturas historicas que tengan estas etiquetas.
        # Deriva la conversacion via Customer → Lead (telefono match) y toma
        # la conversacion mas reciente del Lead.
        historico = frappe.db.sql("""
            SELECT
                ela.etiqueta AS etiqueta,
                (
                    SELECT MAX(cc.creation)
                    FROM `tabConversacion Chatwoot` cc
                    JOIN `tabLead` L ON L.name = cc.lead
                    JOIN `tabCustomer` C ON C.name = si.customer
                    WHERE L.mobile_no = C.custom_numero OR L.mobile_no = C.mobile_no
                ) AS fecha_contacto
            FROM `tabSales Invoice` si
            JOIN `tabSales Invoice Item` sii ON sii.parent = si.name
            JOIN `tabEtiqueta Lead Asociado` ela
                ON ela.parent = sii.sales_order AND ela.parenttype = 'Sales Order'
            WHERE si.docstatus = 1
              AND ela.etiqueta IN %(etiquetas)s
            GROUP BY si.name, ela.etiqueta
        """, {"etiquetas": etiquetas_relevantes}, as_dict=True)

        for h in historico:
            fc = str(h.fecha_contacto)[:10] if h.fecha_contacto else ""
            if h.etiqueta and fc and (h.etiqueta, fc) in etiq_fecha_con_gasto:
                conteo_etiq_fecha[(h.etiqueta, fc)] += 1

    # Fallback: si no se encontro conteo historico, usar conteo local
    for fp in filas_pre:
        key = (fp["etiqueta_usada"], fp["fecha_contacto"])
        if fp["etiqueta_usada"] and fp["fecha_contacto"] and fp["gasto_dia_total"] > 0:
            if key not in conteo_etiq_fecha:
                conteo_etiq_fecha[key] = 1

    # 12. Ensamblar resultado con gasto prorrateado
    result = []
    for fp in filas_pre:
        r = fp["row"]
        so = fp["so"]

        gasto_dia_total = fp["gasto_dia_total"]
        n_ventas = conteo_etiq_fecha.get((fp["etiqueta_usada"], fp["fecha_contacto"]), 1)
        gasto_prorrateado = round(gasto_dia_total / n_ventas, 2) if gasto_dia_total > 0 else 0

        costo_productos = flt(costo_map.get(so, 0)) if so else 0
        envio = flt(envio_map.get(so, 0)) if so else 0
        ticket = flt(r.grand_total)

        # ROAS con gasto prorrateado
        roas = 0
        if gasto_prorrateado > 0 and ticket:
            roas = round(ticket / gasto_prorrateado, 2)

        # Ganancia real con gasto prorrateado
        ganancia = round(ticket - costo_productos - envio - gasto_prorrateado, 2)

        result.append({
            "vendedora": user_map.get(r.so_owner, r.so_owner or ""),
            "fecha_venta": str(r.posting_date) if r.posting_date else "",
            "fecha_contacto": fp["fecha_contacto"],
            "whatsapp_cliente": phone_map.get(r.customer, "") or "",
            "whatsapp_empresa": fp["whatsapp_empresa"],
            "nombre_campana": fp["campana_nombre"],
            "nombre_conjunto": fp["conjunto_nombre"],
            "nombre_anuncio": fp["anuncio_nombre"],
            "gasto_dia": gasto_prorrateado,
            "gasto_dia_total": gasto_dia_total,
            "n_ventas_anuncio": n_ventas,
            "codigo_campana": fp["codigo_campana"],
            "roas": roas,
            "ticket_venta": ticket,
            "costo_producto": costo_productos,
            "envio": envio,
            "ganancia": ganancia,
            "cliente": r.customer_name or "",
            "invoice": r.invoice_name or ""
        })

    return result


def _calc_costs_per_so(so_names):
    """Calcula costo de productos y envio por Sales Order."""
    costo_map = {}  # so_name -> costo total productos
    envio_map = {}  # so_name -> costo envio (items "Item Extra")

    if not so_names:
        return costo_map, envio_map

    ITEM_EXTRA_CODE = "Item Extra"

    # Items de cada SO
    so_items = frappe.db.sql("""
        SELECT parent, item_code, item_name, qty, rate
        FROM `tabSales Order Item`
        WHERE parent IN %(so_names)s
        ORDER BY parent, idx
    """, {"so_names": so_names}, as_dict=True)

    if not so_items:
        return costo_map, envio_map

    # Item codes de productos reales (no extras)
    product_codes = list(set(
        it.item_code for it in so_items if it.item_code != ITEM_EXTRA_CODE
    ))

    # Costo base desde Item Price (Compra estandar)
    cost_base = {}
    if product_codes:
        prices = frappe.db.sql("""
            SELECT item_code, price_list_rate
            FROM `tabItem Price`
            WHERE price_list = 'Compra estandar'
              AND item_code IN %(codes)s
        """, {"codes": product_codes}, as_dict=True)
        cost_base = {p.item_code: flt(p.price_list_rate) for p in prices}

    # Costos adicionales por item (DocType opcional; en algunos tenants no existe).
    cost_extra = {}
    if product_codes and frappe.db.exists("DocType", "Costo adicional asociado"):
        extras = frappe.db.sql("""
            SELECT parent AS item_code, SUM(valor) AS total_extra
            FROM `tabCosto adicional asociado`
            WHERE parent IN %(codes)s AND parenttype = 'Item'
            GROUP BY parent
        """, {"codes": product_codes}, as_dict=True)
        cost_extra = {e.item_code: flt(e.total_extra) for e in extras}

    # Calcular por SO
    for it in so_items:
        so = it.parent
        monto = flt(it.qty) * flt(it.rate)
        if it.item_code == ITEM_EXTRA_CODE:
            # Solo es envio si el nombre contiene "envio" o "envío"
            nombre = (it.item_name or "").lower()
            if "envio" in nombre or "envío" in nombre:
                envio_map[so] = flt(envio_map.get(so, 0)) + monto
            else:
                # Item Extra que no es envio (ej: grabado, producto adicional) → costo producto
                costo_map[so] = flt(costo_map.get(so, 0)) + monto
        else:
            # Producto real: costo unitario = base + adicionales
            costo_unit = flt(cost_base.get(it.item_code, 0)) + flt(cost_extra.get(it.item_code, 0))
            costo_map[so] = flt(costo_map.get(so, 0)) + flt(it.qty) * costo_unit

    return costo_map, envio_map


def _resolve_chain_from_anuncio(anuncio_name):
    """Construye la cadena Anuncio → Conjunto → Campana partiendo de un Anuncio
    directo (atribucion CTWA via Sales Invoice/Order.custom_anuncio).

    Retorna dict con campana / conjunto_nombre / anuncio / anuncio_meta_id
    o None si el Anuncio no esta enlazado a un Conjunto/Campana.
    """
    if not anuncio_name:
        return None
    anuncio = frappe.db.get_value(
        "Anuncio", anuncio_name, ["name", "nombre", "meta_id"], as_dict=True
    )
    if not anuncio:
        return None

    conjunto_link = frappe.db.get_value(
        "Anuncios Asociados", {"anuncios": anuncio.name}, ["parent"], as_dict=True
    )
    conjunto_nombre_amigable = ""
    campana_nombre = ""
    if conjunto_link:
        conjunto_nombre_amigable = frappe.db.get_value(
            "Conjunto Anuncios", conjunto_link.parent, "nombre"
        ) or conjunto_link.parent

        campana_link = frappe.db.get_value(
            "Conjunto Anuncios Asociados",
            {"conjunto_anuncios": conjunto_link.parent},
            ["parent"],
            as_dict=True,
        )
        if campana_link:
            campana_nombre = frappe.db.get_value(
                "Campana Meta", campana_link.parent, "nombre"
            ) or campana_link.parent

    return {
        "campana": campana_nombre,
        "conjunto_nombre": conjunto_nombre_amigable,
        "anuncio": anuncio.nombre or anuncio.name,
        "anuncio_meta_id": anuncio.meta_id or "",
    }


def _build_campaign_chain(etiquetas):
    """Construye cadena: etiqueta -> Anuncio -> Conjunto -> Campana Meta."""
    if not etiquetas:
        return {}

    # Anuncios que matchean etiquetas
    anuncios = frappe.get_all("Anuncio",
        filters={"etiqueta": ["in", list(etiquetas)]},
        fields=["name", "nombre", "etiqueta", "meta_id"])
    if not anuncios:
        return {}

    # Agrupar anuncios por etiqueta (puede haber varios con la misma etiqueta)
    from collections import defaultdict
    etiqueta_to_anuncios = defaultdict(list)
    for a in anuncios:
        etiqueta_to_anuncios[a.etiqueta].append(a)
    anuncio_names = [a.name for a in anuncios]

    # Conjunto que contiene cada Anuncio
    conjunto_links = frappe.get_all("Anuncios Asociados",
        filters={"anuncios": ["in", anuncio_names]},
        fields=["anuncios", "parent"])
    if not conjunto_links:
        return {}

    anuncio_to_conjunto = {cl.anuncios: cl.parent for cl in conjunto_links}
    conjunto_names = list(set(cl.parent for cl in conjunto_links))

    # Nombres amigables de conjuntos
    conjuntos = frappe.get_all("Conjunto Anuncios",
        filters={"name": ["in", conjunto_names]},
        fields=["name", "nombre"])
    conjunto_nombre_map = {c.name: c.nombre for c in conjuntos}

    # Campana que contiene cada Conjunto
    campana_links = frappe.get_all("Conjunto Anuncios Asociados",
        filters={"conjunto_anuncios": ["in", conjunto_names]},
        fields=["conjunto_anuncios", "parent"])
    if not campana_links:
        return {}

    conjunto_to_campana = {cl.conjunto_anuncios: cl.parent for cl in campana_links}
    campana_names = list(set(cl.parent for cl in campana_links))

    campanas = frappe.get_all("Campana Meta",
        filters={"name": ["in", campana_names]},
        fields=["name", "nombre", "meta_id"])
    campana_map = {c.name: c for c in campanas}

    # Armar cadena (con todos los meta_ids posibles por etiqueta)
    chain = {}
    for etiqueta_name in etiquetas:
        anuncio_list = etiqueta_to_anuncios.get(etiqueta_name, [])
        if not anuncio_list:
            continue

        # Buscar el primer anuncio que tenga cadena completa
        best = None
        all_meta_ids = []
        for anuncio in anuncio_list:
            if anuncio.meta_id:
                all_meta_ids.append(anuncio.meta_id)
            conjunto_name = anuncio_to_conjunto.get(anuncio.name)
            if not conjunto_name:
                continue
            campana_name = conjunto_to_campana.get(conjunto_name)
            if not campana_name:
                continue
            campana = campana_map.get(campana_name)
            if not campana:
                continue
            if not best:
                best = {
                    "anuncio": anuncio.nombre,
                    "anuncio_meta_id": anuncio.meta_id or "",
                    "conjunto_nombre": conjunto_nombre_map.get(conjunto_name, conjunto_name),
                    "campana": campana.nombre,
                    "campaign_meta_id": campana.meta_id
                }

        if best:
            # Guardar todos los meta_ids para probar al buscar gasto
            best["all_meta_ids"] = list(set(mid for mid in all_meta_ids if mid))
            chain[etiqueta_name] = best

    return chain


def _fetch_daily_spend(account_id, from_date, to_date, ad_meta_ids):
    """Obtiene gasto diario por anuncio desde Meta API."""
    token = _conf_get_meta_token()
    if not token:
        return {}

    result = {}
    try:
        ad_filter = json.dumps([
            {"field": "ad.id", "operator": "IN", "value": list(ad_meta_ids)}
        ])

        url = f"https://graph.facebook.com/v22.0/act_{account_id}/insights"
        params = {
            "access_token": token,
            "level": "ad",
            "time_increment": 1,
            "fields": "ad_id,spend",
            "time_range": json.dumps({"since": str(from_date), "until": str(to_date)}),
            "filtering": ad_filter,
            "limit": 500
        }

        while url:
            resp = requests.get(url, params=params)
            if resp.status_code != 200:
                frappe.log_error(f"Meta Spend API Error: {resp.text}", "Meta-Spend-Error")
                break

            data = resp.json()
            for row in data.get("data", []):
                aid = row.get("ad_id", "")
                spend = float(row.get("spend", 0))
                date_start = row.get("date_start", "")
                result.setdefault(aid, {})[date_start] = spend

            url = data.get("paging", {}).get("next")
            params = {}  # next URL ya incluye params

    except Exception as e:
        frappe.log_error(f"Meta Spend Fetch Error: {str(e)}", "Meta-Spend-Error")

    return result


@frappe.whitelist()
def export_ventas_xlsx(from_date=None, to_date=None, account_id=None):
    """Exporta la base de datos de ventas a XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = get_ventas_data(from_date, to_date, account_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Base de Datos"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    money_fmt = '#,##0.00'

    headers = [
        "VENDEDORA", "FECHA DE VENTA", "FECHA DE CONTACTO",
        "WHATSAPP CLIENTE", "WHATSAPP EMPRESA", "NOMBRE DE CAMPANA",
        "CONJUNTO DE ANUNCIOS", "ANUNCIO",
        "GASTO DIA ANUNCIO", "CODIGO DE CAMPANA", "ROAS",
        "TICKET DE VENTA", "COSTO PRODUCTO", "ENVIO", "GANANCIA", "CLIENTE"
    ]
    col_widths = [22, 16, 16, 18, 18, 35, 30, 30, 20, 30, 12, 16, 16, 12, 16, 35]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, r in enumerate(data, 2):
        values = [
            r.get("vendedora", ""),
            r.get("fecha_venta", ""),
            r.get("fecha_contacto", ""),
            r.get("whatsapp_cliente", ""),
            r.get("whatsapp_empresa", ""),
            r.get("nombre_campana", ""),
            r.get("nombre_conjunto", ""),
            r.get("nombre_anuncio", ""),
            r.get("gasto_dia", 0),
            r.get("codigo_campana", ""),
            r.get("roas", 0),
            r.get("ticket_venta", 0),
            r.get("costo_producto", 0),
            r.get("envio", 0),
            r.get("ganancia", 0),
            r.get("cliente", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in (9, 12, 13, 14, 15):
                cell.number_format = money_fmt
            elif col_idx == 11:
                cell.number_format = '0.00'
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    file_label = f"Base_Datos_{from_date or 'all'}_{to_date or 'all'}"
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response["filename"] = f"{file_label}.xlsx"
    frappe.response["filecontent"] = output.getvalue()
    frappe.response["type"] = "binary"


@frappe.whitelist()
def debug_ventas_etiqueta(etiqueta=None, fecha_contacto=None):
    """Debug: muestra todas las ventas historicas de una etiqueta+fecha."""
    if not etiqueta:
        return {"error": "falta parametro etiqueta"}

    historico = frappe.db.sql("""
        SELECT
            si.name AS invoice,
            si.posting_date AS fecha_venta,
            si.customer_name AS cliente,
            si.grand_total AS ticket,
            ela.etiqueta AS etiqueta,
            sii.sales_order,
            (
                SELECT MAX(cc.creation)
                FROM `tabConversacion Chatwoot` cc
                JOIN `tabLead` L ON L.name = cc.lead
                JOIN `tabCustomer` C ON C.name = si.customer
                WHERE L.mobile_no = C.custom_numero OR L.mobile_no = C.mobile_no
            ) AS fecha_contacto
        FROM `tabSales Invoice` si
        JOIN `tabSales Invoice Item` sii ON sii.parent = si.name
        JOIN `tabEtiqueta Lead Asociado` ela
            ON ela.parent = sii.sales_order AND ela.parenttype = 'Sales Order'
        WHERE si.docstatus = 1
          AND ela.etiqueta = %(etiqueta)s
        GROUP BY si.name, ela.etiqueta
        ORDER BY si.posting_date DESC
    """, {"etiqueta": etiqueta}, as_dict=True)

    result = []
    for h in historico:
        fc = str(h.fecha_contacto)[:10] if h.fecha_contacto else ""
        if fecha_contacto and fc != fecha_contacto:
            continue
        result.append({
            "invoice": h.invoice,
            "fecha_venta": str(h.fecha_venta),
            "fecha_contacto": fc,
            "cliente": h.customer_name if hasattr(h, 'customer_name') else h.cliente,
            "ticket": float(h.ticket),
            "sales_order": h.sales_order
        })

    return {"etiqueta": etiqueta, "fecha_filtro": fecha_contacto or "todas", "total": len(result), "ventas": result}
